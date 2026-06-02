"""Master cashbook upload, regional assignment, bank correction, and feed to subsidiaries."""

import re
from datetime import datetime

import pandas as pd

from client_region_service import SUBSIDIARY_LABELS

BILLING_ENTITY_TO_SUBSIDIARY = {
    'Ndevor Systems Ltd : Phorest Australia': 1,
    'Ndevor Systems Ltd : Phorest Canada': 2,
    'Ndevor Systems Ltd : Phorest US': 3,
    'Ndevor Systems Ltd : Phorest Ireland': 4,
    'Ndevor Systems Ltd : Phorest Ireland : Phorest UK': 5,
    'Ndevor Systems Ltd : Phorest Germany': 4,
}

BILLING_ENTITIES = {
    1: 'Ndevor Systems Ltd : Phorest Australia',
    2: 'Ndevor Systems Ltd : Phorest Canada',
    3: 'Ndevor Systems Ltd : Phorest US',
    4: 'Ndevor Systems Ltd : Phorest Ireland',
    5: 'Ndevor Systems Ltd : Phorest Ireland : Phorest UK',
}

BANK_ACCOUNTS = {
    1: '10130 Bank : CB current a/c AU$ # 411110236694',
    2: '10150 Bank : CIBC Current Account 9066314',
    3: '10043 Bank : CIBC operating a/c US$ # 2605090',
    4: '10010 Bank : BOI current a/c EUR # 17013705',
    5: '10020 Bank : BOI current a/c GBP # 62100285',
}

BANK_ACCOUNT_TO_SUBSIDIARY = {v: k for k, v in BANK_ACCOUNTS.items()}
BANK_ACCOUNT_TO_SUBSIDIARY[
    '10010c Bank : Dummy Interco Bank Accounts : Interco - BOI current a/c Ä # 17013705 (Germany)'
] = 4

GERMANY_BILLING = 'Ndevor Systems Ltd : Phorest Germany'
GERMANY_BANK = '10010c Bank : Dummy Interco Bank Accounts : Interco - BOI current a/c Ä # 17013705 (Germany)'
GERMANY_LOCATIONS = {'Germany', 'Switzerland', 'Austria'}

CURRENCY_TO_SUBSIDIARY = {
    'AUD': 1,
    'CAD': 2,
    'USD': 3,
    'EUR': 4,
    'AED': 4,
    'GBP': 5,
}

SEPA_TRANSTYPES = {'SEPASUCCESS', 'SEPAREJECT'}

DATA_ERROR_BILLING = 'DATA ERROR'

# Currency hints for DATA ERROR rows (same rules as Looker fix-errors)
DATA_ERROR_CURRENCY_TO_SUBSIDIARY = {
    'CAD': 2,
    'AED': 4,
}

# Weak location → Stripe region hints when billing entity is DATA ERROR
LOCATION_TO_SUBSIDIARY = {
    'australia': 1,
    'canada': 2,
    'united states': 3,
    'usa': 3,
    'us': 3,
    'ireland': 4,
    'germany': 4,
    'switzerland': 4,
    'austria': 4,
    'france': 4,
    'spain': 4,
    'italy': 4,
    'netherlands': 4,
    'belgium': 4,
    'united kingdom': 5,
    'uk': 5,
}

COLUMN_ALIASES = {
    'payment_date': ['payment_date', 'Payment Date', 'payment date'],
    'client_id': ['client_id', 'Client ID', 'client id'],
    'invoice_number': ['invoice_number', 'Invoice Number', 'invoice number'],
    'billing_entity': ['billing_entity', 'Billing Entity', 'billing entity'],
    'ar_account': ['ar_account', 'AR Account', 'ar account'],
    'currency': ['currency', 'Currency'],
    'exchange_rate': ['exchange_rate', 'Exchange Rate', 'exchange rate'],
    'amount': ['amount', 'Amount'],
    'account': ['account', 'Account'],
    'location': ['location', 'Location'],
    'transtype': ['transtype', 'Transtype', 'Transtype'],
    'comment': ['comment', 'Comment'],
    'card_reference': ['card_reference', 'Card Reference', 'card reference'],
    'reasoncode': ['reasoncode', 'Reasoncode', 'Reasoncode'],
    'sepaprovider': ['sepaprovider', 'SEPA Provider', 'sepa provider'],
    'invoice_hash': ['invoice_hash', 'invoice #', 'invoice#'],
    'payment_hash': ['payment_hash', 'payment #', 'payment#'],
    'memo': ['memo', 'Memo'],
}


def _normalize(name):
    return re.sub(r'[^a-z0-9]', '', str(name).lower())


def normalize_cashbook_columns(df):
    """Rename DataFrame columns to canonical snake_case field names."""
    colmap = {_normalize(c): c for c in df.columns}
    rename = {}
    for canonical, candidates in COLUMN_ALIASES.items():
        for cand in candidates:
            key = _normalize(cand)
            if key in colmap:
                rename[colmap[key]] = canonical
                break
    out = df.rename(columns=rename)
    return out


def _safe_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    return str(val).strip()


def _safe_float(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _payment_date_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        s = val.strip()
        if s:
            return s
        return None
    try:
        if hasattr(val, 'strftime'):
            return val.strftime('%d/%m/%Y')
    except Exception:
        pass
    return str(val)


def _is_germany_row(location, billing_entity):
    loc = _safe_str(location)
    if loc in GERMANY_LOCATIONS:
        return True
    return _safe_str(billing_entity) == GERMANY_BILLING


def is_data_error_billing(billing_entity):
    return _safe_str(billing_entity).upper() == DATA_ERROR_BILLING


def subsidiary_from_billing_entity(billing_entity):
    be = _safe_str(billing_entity)
    if not be or is_data_error_billing(be):
        return None
    return BILLING_ENTITY_TO_SUBSIDIARY.get(be)


def subsidiary_from_data_error_currency(currency):
    """Resolve DATA ERROR rows using currency (Looker CAD/AED pattern)."""
    c = _safe_str(currency).upper()
    return DATA_ERROR_CURRENCY_TO_SUBSIDIARY.get(c)


def subsidiary_from_location(location):
    """Weak region hint from cashbook Location when billing entity is unusable."""
    loc = _safe_str(location).lower()
    if not loc:
        return None
    if loc in LOCATION_TO_SUBSIDIARY:
        return LOCATION_TO_SUBSIDIARY[loc]
    for key, sid in LOCATION_TO_SUBSIDIARY.items():
        if key in loc:
            return sid
    return None


def subsidiary_from_account(account):
    ac = _safe_str(account)
    if not ac:
        return None
    return BANK_ACCOUNT_TO_SUBSIDIARY.get(ac)


def subsidiary_from_currency(currency):
    c = _safe_str(currency).upper()
    return CURRENCY_TO_SUBSIDIARY.get(c)


def canonical_for_subsidiary(subsidiary_id, location=None, billing_entity=None):
    """Return (billing_entity, account) for assigned subsidiary, with Germany exception."""
    if subsidiary_id == 4 and _is_germany_row(location, billing_entity):
        return GERMANY_BILLING, GERMANY_BANK
    return BILLING_ENTITIES.get(subsidiary_id, ''), BANK_ACCOUNTS.get(subsidiary_id, '')


def assign_row_subsidiary(row, profile_by_client):
    """Return (subsidiary_id, source, detail)."""
    client_id = row.client_id
    client_key = str(client_id).strip() if client_id is not None else ''
    profile = profile_by_client.get(client_key) if client_key else None
    data_error = is_data_error_billing(row.billing_entity)

    if profile and not profile.is_multi_region and profile.primary_subsidiary_id:
        detail = 'DATA ERROR billing; assigned from client region map' if data_error else None
        return profile.primary_subsidiary_id, 'client_mapping', detail

    if profile and profile.is_multi_region:
        sid = subsidiary_from_billing_entity(row.billing_entity)
        if sid:
            return sid, 'multi_region_row_billing', None
        sid = subsidiary_from_account(row.account)
        if not sid and data_error:
            sid = subsidiary_from_data_error_currency(row.currency)
        if not sid:
            sid = subsidiary_from_currency(row.currency)
        if not sid and data_error:
            sid = subsidiary_from_location(row.location)
        if sid:
            return sid, 'multi_region_row_billing', (
                'DATA ERROR billing; used row clues (not billing entity)' if data_error else
                'Used row clues; billing entity unknown'
            )
        if profile.primary_subsidiary_id:
            return (
                profile.primary_subsidiary_id,
                'multi_region_row_billing',
                'DATA ERROR billing; used primary region from map' if data_error else
                'Used primary region; row billing entity unknown',
            )

    if not data_error:
        sid = subsidiary_from_billing_entity(row.billing_entity)
        if sid:
            return sid, 'billing_entity', None

    sid = subsidiary_from_account(row.account)
    if sid:
        detail = 'DATA ERROR billing; resolved via bank account' if data_error else None
        return sid, 'account', detail

    if data_error:
        sid = subsidiary_from_data_error_currency(row.currency)
        if sid:
            return sid, 'data_error_currency', 'DATA ERROR billing; resolved via currency (CAD/AED)'

    sid = subsidiary_from_currency(row.currency)
    if sid:
        detail = 'DATA ERROR billing; resolved via currency' if data_error else None
        return sid, 'currency', detail

    if data_error:
        sid = subsidiary_from_location(row.location)
        if sid:
            return sid, 'location', 'DATA ERROR billing; resolved via Location'

    tt = _safe_str(row.transtype).upper()
    if tt in SEPA_TRANSTYPES or _safe_str(row.sepaprovider):
        detail = 'DATA ERROR billing; resolved via SEPA transtype' if data_error else None
        return 4, 'transtype', detail

    if data_error:
        return None, 'unresolved', (
            'DATA ERROR billing — no region clue; try matching in each Stripe region during reconciliation'
        )

    return None, 'unresolved', 'Could not determine region'


def apply_field_corrections(row):
    """Rewrite billing_entity/account when assignment warrants it. Returns True if corrected."""
    sid = row.assigned_subsidiary_id
    if not sid or row.assignment_source == 'unresolved':
        return False

    canon_be, canon_ac = canonical_for_subsidiary(
        sid, location=row.location, billing_entity=row.original_billing_entity or row.billing_entity
    )
    if not canon_be and not canon_ac:
        return False

    had_data_error = is_data_error_billing(row.original_billing_entity)
    force = (
        row.assignment_source in ('client_mapping', 'multi_region_row_billing', 'data_error_currency')
        or had_data_error
    )
    changed = False

    if force:
        if row.billing_entity != canon_be and canon_be:
            row.billing_entity = canon_be
            changed = True
        if row.account != canon_ac and canon_ac:
            row.account = canon_ac
            changed = True
    else:
        expected_ac = BANK_ACCOUNTS.get(sid) if sid != 4 or not _is_germany_row(row.location, row.billing_entity) else GERMANY_BANK
        if expected_ac and row.account != expected_ac:
            row.account = expected_ac
            changed = True
        if not _safe_str(row.billing_entity) and canon_be:
            row.billing_entity = canon_be
            changed = True

    if changed:
        row.fields_corrected = True
    return changed


def row_from_series(job_id, filename, series):
    """Build kwargs dict for MasterCashbookTransaction from a pandas Series."""
    return {
        'job_id': job_id,
        'payment_date': _payment_date_str(series.get('payment_date')),
        'client_id': int(series['client_id']) if pd.notna(series.get('client_id')) else None,
        'invoice_number': _safe_str(series.get('invoice_number')),
        'billing_entity': _safe_str(series.get('billing_entity')),
        'ar_account': _safe_str(series.get('ar_account')),
        'currency': _safe_str(series.get('currency')),
        'exchange_rate': _safe_float(series.get('exchange_rate')),
        'amount': _safe_float(series.get('amount')),
        'account': _safe_str(series.get('account')),
        'location': _safe_str(series.get('location')),
        'transtype': _safe_str(series.get('transtype')),
        'comment': _safe_str(series.get('comment')) or None,
        'card_reference': _safe_float(series.get('card_reference')),
        'reasoncode': _safe_float(series.get('reasoncode')),
        'sepaprovider': _safe_str(series.get('sepaprovider')) or None,
        'invoice_hash': _safe_str(series.get('invoice_hash')),
        'payment_hash': _safe_str(series.get('payment_hash')),
        'memo': _safe_float(series.get('memo')),
        'filename': filename,
        'uploaded_at': datetime.utcnow(),
    }


def load_master_from_dataframe(db, MasterCashbookTransaction, ClientRegionProfile, job_id, df, filename):
    """Replace master rows for job, assign regions, apply corrections."""
    df = normalize_cashbook_columns(df)
    if 'client_id' not in df.columns:
        raise ValueError('Missing client_id column')

    skipped_no_client = 0
    profiles = {p.client_id: p for p in ClientRegionProfile.query.all()}

    MasterCashbookTransaction.query.filter_by(job_id=job_id).delete()
    db.session.flush()

    rows_added = 0
    data_error_rows = 0
    data_error_resolved = 0
    data_error_unresolved = 0
    for _, series in df.iterrows():
        if pd.isna(series.get('client_id')):
            skipped_no_client += 1
            continue

        kwargs = row_from_series(job_id, filename, series)
        row = MasterCashbookTransaction(**kwargs)
        row.original_account = row.account
        row.original_billing_entity = row.billing_entity

        sid, source, detail = assign_row_subsidiary(row, profiles)
        row.assigned_subsidiary_id = sid
        row.assignment_source = source
        row.assignment_detail = detail

        apply_field_corrections(row)
        if is_data_error_billing(row.original_billing_entity):
            if row.assigned_subsidiary_id:
                data_error_resolved += 1
            else:
                data_error_unresolved += 1
        db.session.add(row)
        rows_added += 1

    db.session.commit()
    return {
        'rows_added': rows_added,
        'skipped_no_client': skipped_no_client,
        'data_error_rows': data_error_rows,
        'data_error_resolved': data_error_resolved,
        'data_error_unresolved': data_error_unresolved,
    }


def reassign_master_job(db, MasterCashbookTransaction, ClientRegionProfile, job_id):
    """Re-run assignment and corrections on existing master rows."""
    profiles = {p.client_id: p for p in ClientRegionProfile.query.all()}
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    for row in rows:
        sid, source, detail = assign_row_subsidiary(row, profiles)
        row.assigned_subsidiary_id = sid
        row.assignment_source = source
        row.assignment_detail = detail
        apply_field_corrections(row)
    db.session.commit()
    return {'rows_updated': len(rows)}


def master_status(db, MasterCashbookTransaction, CashbookTransaction, ProcessingJob, job_id):
    """Aggregate stats for API status endpoint."""
    job = ProcessingJob.query.get(job_id)
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    total = len(rows)
    by_region = {label: 0 for label in SUBSIDIARY_LABELS.values()}
    by_region_amount = {label: 0.0 for label in SUBSIDIARY_LABELS.values()}
    by_source = {}
    unresolved = 0
    corrected = 0
    data_error_rows = 0
    data_error_unresolved = 0
    total_amount = 0.0
    unresolved_amount = 0.0
    last_upload = None
    filename = None

    for r in rows:
        amt = float(r.amount or 0)
        total_amount += amt
        if is_data_error_billing(r.original_billing_entity):
            data_error_rows += 1
            if not r.assigned_subsidiary_id:
                data_error_unresolved += 1
        if r.assigned_subsidiary_id:
            label = SUBSIDIARY_LABELS.get(r.assigned_subsidiary_id, '?')
            by_region[label] = by_region.get(label, 0) + 1
            by_region_amount[label] = by_region_amount.get(label, 0.0) + amt
        else:
            unresolved += 1
            unresolved_amount += amt
        src = r.assignment_source or 'unknown'
        by_source[src] = by_source.get(src, 0) + 1
        if r.fields_corrected:
            corrected += 1
        if r.uploaded_at and (last_upload is None or r.uploaded_at > last_upload):
            last_upload = r.uploaded_at
            filename = r.filename

    feed_status = {}
    workbook_by_region = {}
    to_be_uploaded = 0
    for r in rows:
        if is_to_be_uploaded_row(r):
            to_be_uploaded += 1

    for sid, label in SUBSIDIARY_LABELS.items():
        cb_count = CashbookTransaction.query.filter_by(job_id=job_id, subsidiary_id=sid).count()
        master_count = sum(1 for r in rows if r.assigned_subsidiary_id == sid)
        wb_count = sum(1 for r in rows if r.workbook_region_id == sid)
        wb_updates = [r.workbook_updated_at for r in rows if r.workbook_region_id == sid and r.workbook_updated_at]
        feed_status[label] = {
            'subsidiary_id': sid,
            'master_assigned_rows': master_count,
            'cashbook_transaction_rows': cb_count,
            'fed': cb_count > 0,
        }
        workbook_by_region[label] = {
            'subsidiary_id': sid,
            'tab_key': TAB_KEY_FOR_SUBSIDIARY.get(sid),
            'workbook_rows': wb_count,
            'last_updated': max(wb_updates).isoformat() if wb_updates else None,
        }

    return {
        'job_id': job_id,
        'job_name': job.job_name if job else f'Job {job_id}',
        'total_rows': total,
        'total_amount': round(total_amount, 2),
        'unresolved_amount': round(unresolved_amount, 2),
        'unresolved_rows': unresolved,
        'fields_corrected_rows': corrected,
        'by_region': by_region,
        'by_region_amount': {k: round(v, 2) for k, v in by_region_amount.items()},
        'by_assignment_source': by_source,
        'last_upload_at': last_upload.isoformat() if last_upload else None,
        'filename': filename,
        'feed_status': feed_status,
        'data_error_rows': data_error_rows,
        'data_error_unresolved': data_error_unresolved,
        'workbook_summary': {
            'to_be_uploaded': to_be_uploaded,
            'by_region': workbook_by_region,
        },
    }


def preview_issues(MasterCashbookTransaction, job_id, limit=50):
    """Sample unresolved, DATA ERROR, or flagged rows."""
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    flagged = []
    for r in rows:
        if (
            r.assigned_subsidiary_id is None
            or r.assignment_detail
            or is_data_error_billing(r.original_billing_entity)
        ):
            flagged.append(r)
    items = []
    for r in flagged[:limit]:
        items.append({
            'client_id': r.client_id,
            'billing_entity': r.original_billing_entity or r.billing_entity,
            'billing_entity_corrected': r.billing_entity if r.fields_corrected else None,
            'account': (r.account or '')[:80],
            'currency': r.currency,
            'location': r.location,
            'transtype': r.transtype,
            'assigned_region': SUBSIDIARY_LABELS.get(r.assigned_subsidiary_id, '—'),
            'assignment_source': r.assignment_source,
            'assignment_detail': r.assignment_detail,
            'fields_corrected': r.fields_corrected,
            'is_data_error': is_data_error_billing(r.original_billing_entity),
        })
    return items


def master_rows_for_export(MasterCashbookTransaction, job_id, subsidiary_id=None):
    q = MasterCashbookTransaction.query.filter_by(job_id=job_id)
    if subsidiary_id is not None:
        q = q.filter_by(assigned_subsidiary_id=subsidiary_id)
    return q.order_by(MasterCashbookTransaction.id).all()


def master_to_export_dataframe(rows):
    """Export format matching subsidiary cashbook CSV headers."""
    data = []
    for t in rows:
        data.append({
            'payment_date': t.payment_date,
            'client_id': t.client_id,
            'invoice_number': t.invoice_number,
            'billing_entity': t.billing_entity,
            'ar_account': t.ar_account,
            'currency': t.currency,
            'exchange_rate': t.exchange_rate,
            'amount': t.amount,
            'account': t.account,
            'Location': t.location,
            'transtype': t.transtype,
            'comment': t.comment,
            'Card Reference': t.card_reference,
            'reasoncode': t.reasoncode,
            'sepaprovider': t.sepaprovider,
            'invoice #': t.invoice_hash,
            'payment #': t.payment_hash,
            'Memo': t.memo,
        })
    return pd.DataFrame(data)


def feed_subsidiary(db, MasterCashbookTransaction, CashbookTransaction, job_id, subsidiary_id):
    """Copy assigned master rows into cashbook_transactions for one subsidiary."""
    master_rows = MasterCashbookTransaction.query.filter_by(
        job_id=job_id, assigned_subsidiary_id=subsidiary_id
    ).all()

    CashbookTransaction.query.filter_by(job_id=job_id, subsidiary_id=subsidiary_id).delete()
    db.session.flush()

    added = 0
    for m in master_rows:
        if is_separated_workbook_row(m) or is_uk_bacs_transtype(m.transtype) or is_partner_transfer_transtype(m.transtype):
            continue
        tx = CashbookTransaction(
            subsidiary_id=subsidiary_id,
            job_id=job_id,
            master_cashbook_id=m.id,
            payment_date=m.payment_date,
            client_id=m.client_id,
            invoice_number=m.invoice_number,
            billing_entity=m.billing_entity,
            ar_account=m.ar_account,
            currency=m.currency,
            exchange_rate=m.exchange_rate,
            amount=m.amount,
            account=m.account,
            location=m.location,
            transtype=m.transtype,
            comment=m.comment,
            card_reference=m.card_reference,
            reasoncode=m.reasoncode,
            sepaprovider=m.sepaprovider,
            invoice_hash=m.invoice_hash,
            payment_hash=m.payment_hash,
            memo=m.memo,
            filename=m.filename or 'master_cashbook_feed',
            uploaded_at=datetime.utcnow(),
        )
        db.session.add(tx)
        added += 1

    db.session.commit()
    return {'subsidiary_id': subsidiary_id, 'region': SUBSIDIARY_LABELS.get(subsidiary_id), 'rows_fed': added}


def feed_all_subsidiaries(db, MasterCashbookTransaction, CashbookTransaction, job_id):
    results = []
    for sid in SUBSIDIARY_LABELS:
        results.append(feed_subsidiary(db, MasterCashbookTransaction, CashbookTransaction, job_id, sid))
    return results


# ==================== WORKBOOK (Master / To be Uploaded / Region Stripe Import) ====================

WORKBOOK_TAB_KEYS = {
    'master': None,
    'to_be_uploaded': 'to_be_uploaded',
    'uk_bacs': 'uk_bacs',
    'partner_transfer': 'partner_transfer',
    'sepa_netting': 'sepa_netting',
    'au': 1,
    'ca': 2,
    'us': 3,
    'eu': 4,
    'uk': 5,
}

TAB_KEY_FOR_SUBSIDIARY = {v: k for k, v in WORKBOOK_TAB_KEYS.items() if isinstance(v, int)}

UK_BACS_TRANSTYPES = frozenset({'WINSUCCESS', 'UKDDREJECT'})
PARTNER_TRANSFER_TRANSTYPES = frozenset({'PARTNERTRANSFER'})
SEPA_NETTING_TRANSTYPES = frozenset({'SEPASUCCESS', 'SEPAREJECT'})


def is_uk_bacs_transtype(transtype):
    return _safe_str(transtype).upper() in UK_BACS_TRANSTYPES


def is_partner_transfer_transtype(transtype):
    return _safe_str(transtype).upper() in PARTNER_TRANSFER_TRANSTYPES


def is_sepa_netting_transtype(transtype):
    return _safe_str(transtype).upper() in SEPA_NETTING_TRANSTYPES


def _sepa_amount_key(amount):
    if amount is None:
        return None
    try:
        return round(abs(float(amount)), 2)
    except (TypeError, ValueError):
        return None


def find_sepa_netting_pairs(rows):
    """Pair SEPAREJECT with SEPASUCCESS by client_id + abs(amount); prefer invoice match."""
    from collections import defaultdict

    eligible = [r for r in rows if not getattr(r, 'workbook_sepa_netting', False)]
    successes = [r for r in eligible if _safe_str(r.transtype).upper() == 'SEPASUCCESS']
    rejects = [r for r in eligible if _safe_str(r.transtype).upper() == 'SEPAREJECT']

    by_key = defaultdict(list)
    by_key_invoice = defaultdict(list)
    for s in successes:
        cid = _safe_str(s.client_id)
        amt = _sepa_amount_key(s.amount)
        if not cid or amt is None:
            continue
        by_key[(cid, amt)].append(s)
        inv = _safe_str(s.invoice_number)
        if inv:
            by_key_invoice[(cid, amt, inv)].append(s)

    used_success_ids = set()
    pairs = []
    for rej in rejects:
        cid = _safe_str(rej.client_id)
        amt = _sepa_amount_key(rej.amount)
        if not cid or amt is None:
            continue

        inv = _safe_str(rej.invoice_number)
        match = None
        if inv:
            for s in by_key_invoice.get((cid, amt, inv), []):
                if s.id not in used_success_ids:
                    match = s
                    break
        if not match:
            for s in by_key.get((cid, amt), []):
                if s.id not in used_success_ids:
                    match = s
                    break
        if match:
            used_success_ids.add(match.id)
            pairs.append((match, rej))
    return pairs


def count_sepa_netting_candidates(rows):
    return len(find_sepa_netting_pairs(rows))


def is_separated_workbook_row(row):
    """Row moved to a special workbook tab — excluded from To be Uploaded / regional feed."""
    return (
        bool(getattr(row, 'workbook_uk_bacs', False))
        or bool(getattr(row, 'workbook_partner_transfer', False))
        or bool(getattr(row, 'workbook_sepa_netting', False))
    )


def is_to_be_uploaded_row(row):
    """Row still waiting for Stripe Import or special-tab assignment."""
    return row.workbook_region_id is None and not is_separated_workbook_row(row)


def _workbook_to_be_uploaded_query(q, MasterCashbookTransaction):
    """SQL filter matching is_to_be_uploaded_row."""
    from sqlalchemy import or_
    return q.filter(
        MasterCashbookTransaction.workbook_region_id.is_(None),
        or_(
            MasterCashbookTransaction.workbook_uk_bacs.is_(False),
            MasterCashbookTransaction.workbook_uk_bacs.is_(None),
        ),
        or_(
            MasterCashbookTransaction.workbook_partner_transfer.is_(False),
            MasterCashbookTransaction.workbook_partner_transfer.is_(None),
        ),
        or_(
            MasterCashbookTransaction.workbook_sepa_netting.is_(False),
            MasterCashbookTransaction.workbook_sepa_netting.is_(None),
        ),
    )

# Standard Stripe card/SEPA transtypes — anything else matched to Stripe is flagged
NORMAL_STRIPE_MATCH_TRANSTYPES = frozenset({
    'CARDSUCCESS',
    'SEPASUCCESS',
    'SEPAREJECT',
})


def is_unusual_transtype_for_stripe_match(transtype):
    """Return (is_unusual, label). Unusual = not a normal Stripe card/SEPA type."""
    tt = _safe_str(transtype).upper()
    if not tt:
        return True, 'MISSING'
    if tt in NORMAL_STRIPE_MATCH_TRANSTYPES:
        return False, None
    return True, tt


def _apply_unusual_transtype_query(q, MasterCashbookTransaction):
    """Filter query to rows with non-standard transtype (MANUAL, empty, etc.)."""
    from sqlalchemy import func, or_

    tt = func.upper(func.trim(MasterCashbookTransaction.transtype))
    normal = list(NORMAL_STRIPE_MATCH_TRANSTYPES)
    return q.filter(
        or_(
            MasterCashbookTransaction.transtype.is_(None),
            func.trim(MasterCashbookTransaction.transtype) == '',
            ~tt.in_(normal),
        )
    )


def _count_unusual_transtype(rows):
    return sum(1 for r in rows if is_unusual_transtype_for_stripe_match(r.transtype)[0])


def master_row_workbook_dict(row):
    """Serialize a master row for workbook UI/export with unusual-transtype flags."""
    unusual, label = is_unusual_transtype_for_stripe_match(row.transtype)
    matched = row.workbook_region_id is not None
    uk_bacs = bool(getattr(row, 'workbook_uk_bacs', False))
    partner_transfer = bool(getattr(row, 'workbook_partner_transfer', False))
    sepa_netting = bool(getattr(row, 'workbook_sepa_netting', False))
    return {
        **row.to_dict(),
        'unusual_transtype': unusual,
        'unusual_transtype_label': label,
        'matched_in_workbook': matched,
        'highlight_unusual_match': matched and unusual,
        'in_uk_bacs_tab': uk_bacs,
        'is_uk_bacs_transtype': is_uk_bacs_transtype(row.transtype),
        'in_partner_transfer_tab': partner_transfer,
        'is_partner_transfer_transtype': is_partner_transfer_transtype(row.transtype),
        'in_sepa_netting_tab': sepa_netting,
        'is_sepa_netting_transtype': is_sepa_netting_transtype(row.transtype),
    }


def _norm_hash(val):
    if val is None:
        return None
    s = str(val).strip().casefold()
    if not s or s == 'nan':
        return None
    return s


def _reasoncode_key(val):
    """Normalize reasoncode for lookup; return None for empty/sentinel values."""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ('2', '2.0', 'nan'):
            return None
        try:
            val = float(s)
        except ValueError:
            return s
    try:
        f = float(val)
        if f in (2.0, 2):
            return None
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return None


def _build_master_lookup(master_rows):
    """Build lookup indexes for linking matches to master rows."""
    by_id = {r.id: r for r in master_rows}
    by_reasoncode = {}
    by_payment_hash = {}
    for r in master_rows:
        rc = _reasoncode_key(r.reasoncode)
        if rc is not None:
            by_reasoncode[rc] = r.id
        ph = _norm_hash(r.payment_hash)
        if ph:
            by_payment_hash[ph] = r.id
    return by_id, by_reasoncode, by_payment_hash


def _resolve_match_to_master_id(match, CashbookTransaction, by_id, by_reasoncode, by_payment_hash):
    """Return master row id for a matched transaction, or None."""
    cb = CashbookTransaction.query.get(match.cashbook_id)
    if cb and cb.master_cashbook_id and cb.master_cashbook_id in by_id:
        return cb.master_cashbook_id

    rc = _reasoncode_key(match.cb_reasoncode)
    if rc is not None and rc in by_reasoncode:
        return by_reasoncode[rc]

    ph = _norm_hash(match.cb_payment_hash)
    if ph and ph in by_payment_hash:
        return by_payment_hash[ph]

    # Composite fallback
    if match.cb_client_id is not None and ph:
        for mid, row in by_id.items():
            if row.client_id == match.cb_client_id and _norm_hash(row.payment_hash) == ph:
                return mid
    if match.cb_client_id is not None and match.cb_amount is not None and match.cb_payment_date:
        for mid, row in by_id.items():
            if (
                row.client_id == match.cb_client_id
                and row.amount is not None
                and abs(float(row.amount) - float(match.cb_amount)) < 0.01
                and _safe_str(row.payment_date) == _safe_str(match.cb_payment_date)
            ):
                return mid
    return None


def resolve_matched_master_ids(job_id, subsidiary_id, MatchedTransaction, CashbookTransaction, MasterCashbookTransaction):
    """Resolve matched transactions to master row IDs for one region."""
    master_rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    by_id, by_reasoncode, by_payment_hash = _build_master_lookup(master_rows)

    matches = MatchedTransaction.query.filter_by(job_id=job_id, subsidiary_id=subsidiary_id).all()
    master_ids = set()
    stats = {'linked_by_fk': 0, 'linked_by_reasoncode': 0, 'linked_by_payment_hash': 0, 'linked_by_composite': 0, 'unresolved': 0}

    for match in matches:
        cb = CashbookTransaction.query.get(match.cashbook_id)
        if cb and cb.master_cashbook_id and cb.master_cashbook_id in by_id:
            master_ids.add(cb.master_cashbook_id)
            stats['linked_by_fk'] += 1
            continue

        rc = _reasoncode_key(match.cb_reasoncode)
        if rc is not None and rc in by_reasoncode:
            master_ids.add(by_reasoncode[rc])
            stats['linked_by_reasoncode'] += 1
            continue

        ph = _norm_hash(match.cb_payment_hash)
        if ph and ph in by_payment_hash:
            master_ids.add(by_payment_hash[ph])
            stats['linked_by_payment_hash'] += 1
            continue

        mid = _resolve_match_to_master_id(match, CashbookTransaction, by_id, by_reasoncode, by_payment_hash)
        if mid:
            master_ids.add(mid)
            stats['linked_by_composite'] += 1
        else:
            stats['unresolved'] += 1

    return master_ids, stats


def sync_workbook_from_region(db, job_id, subsidiary_id, MatchedTransaction, CashbookTransaction, MasterCashbookTransaction):
    """Update workbook region tab from current Stripe matches."""
    if subsidiary_id not in SUBSIDIARY_LABELS:
        raise ValueError('Invalid subsidiary')

    master_count = MasterCashbookTransaction.query.filter_by(job_id=job_id).count()
    if master_count == 0:
        raise ValueError('No master cashbook loaded for this job')

    now = datetime.utcnow()
    matched_ids, link_stats = resolve_matched_master_ids(
        job_id, subsidiary_id, MatchedTransaction, CashbookTransaction, MasterCashbookTransaction
    )

    rows_moved = 0
    rows_removed = 0

    for row in MasterCashbookTransaction.query.filter_by(job_id=job_id).all():
        if is_separated_workbook_row(row):
            continue
        if row.id in matched_ids:
            if row.workbook_region_id != subsidiary_id:
                row.workbook_region_id = subsidiary_id
                row.workbook_updated_at = now
                rows_moved += 1
            else:
                row.workbook_updated_at = now
        elif row.workbook_region_id == subsidiary_id:
            row.workbook_region_id = None
            row.workbook_updated_at = now
            rows_removed += 1

    db.session.commit()

    to_be_uploaded = sum(
        1 for r in MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
        if is_to_be_uploaded_row(r)
    )
    region_label = SUBSIDIARY_LABELS.get(subsidiary_id)

    return {
        'subsidiary_id': subsidiary_id,
        'region': region_label,
        'rows_moved': rows_moved,
        'rows_in_region_tab': len(matched_ids),
        'rows_removed': rows_removed,
        'to_be_uploaded_remaining': to_be_uploaded,
        'unresolved': link_stats['unresolved'],
        'link_stats': link_stats,
        'master_cashbook_url': f'/prepare/master-cashbook/{job_id}',
    }


def sync_workbook_all_regions(db, job_id, MatchedTransaction, CashbookTransaction, MasterCashbookTransaction):
    """Run sync_workbook_from_region for each subsidiary that has matches."""
    results = []
    for sid in SUBSIDIARY_LABELS:
        match_count = MatchedTransaction.query.filter_by(job_id=job_id, subsidiary_id=sid).count()
        if match_count == 0:
            continue
        results.append(sync_workbook_from_region(
            db, job_id, sid, MatchedTransaction, CashbookTransaction, MasterCashbookTransaction
        ))
    to_be_uploaded = sum(
        1 for r in MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
        if is_to_be_uploaded_row(r)
    )
    return {'regions_updated': len(results), 'results': results, 'to_be_uploaded_remaining': to_be_uploaded}


def clear_workbook_region(db, job_id, subsidiary_id, MasterCashbookTransaction):
    """Clear workbook placement for a region; rows return to To be Uploaded."""
    if subsidiary_id not in SUBSIDIARY_LABELS:
        raise ValueError('Invalid subsidiary')

    rows = MasterCashbookTransaction.query.filter_by(
        job_id=job_id, workbook_region_id=subsidiary_id
    ).all()
    now = datetime.utcnow()
    for row in rows:
        row.workbook_region_id = None
        row.workbook_updated_at = now

    db.session.commit()
    to_be_uploaded = sum(
        1 for r in MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
        if is_to_be_uploaded_row(r)
    )
    return {
        'subsidiary_id': subsidiary_id,
        'region': SUBSIDIARY_LABELS.get(subsidiary_id),
        'rows_cleared': len(rows),
        'to_be_uploaded_remaining': to_be_uploaded,
        'master_cashbook_url': f'/prepare/master-cashbook/{job_id}',
    }


def check_uk_bacs_transtypes(db, job_id, MasterCashbookTransaction):
    """Find WINSUCCESS / UKDDREJECT rows and move them to the UK Bacs workbook tab."""
    master_count = MasterCashbookTransaction.query.filter_by(job_id=job_id).count()
    if master_count == 0:
        raise ValueError('No master cashbook loaded for this job')

    now = datetime.utcnow()
    rows_moved = 0
    already_in_tab = 0
    by_transtype = {tt: 0 for tt in UK_BACS_TRANSTYPES}
    removed_from_region = 0

    for row in MasterCashbookTransaction.query.filter_by(job_id=job_id).all():
        if not is_uk_bacs_transtype(row.transtype):
            continue
        tt = _safe_str(row.transtype).upper()
        if tt in by_transtype:
            by_transtype[tt] += 1

        if row.workbook_uk_bacs:
            already_in_tab += 1
            continue

        if row.workbook_region_id is not None:
            row.workbook_region_id = None
            removed_from_region += 1
        row.workbook_uk_bacs = True
        row.workbook_updated_at = now
        rows_moved += 1

    db.session.commit()

    all_rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    uk_bacs_total = sum(1 for r in all_rows if r.workbook_uk_bacs)
    candidates = sum(1 for r in all_rows if is_uk_bacs_transtype(r.transtype) and not r.workbook_uk_bacs)
    to_be_uploaded = sum(1 for r in all_rows if is_to_be_uploaded_row(r))

    return {
        'rows_moved': rows_moved,
        'already_in_tab': already_in_tab,
        'uk_bacs_total': uk_bacs_total,
        'uk_bacs_candidates_remaining': candidates,
        'removed_from_region_tab': removed_from_region,
        'by_transtype': by_transtype,
        'to_be_uploaded_remaining': to_be_uploaded,
        'master_cashbook_url': f'/prepare/master-cashbook/{job_id}',
    }


def check_partner_transfer_transtypes(db, job_id, MasterCashbookTransaction):
    """Find PARTNERTRANSFER rows and move them to the Partner Transfer workbook tab."""
    master_count = MasterCashbookTransaction.query.filter_by(job_id=job_id).count()
    if master_count == 0:
        raise ValueError('No master cashbook loaded for this job')

    now = datetime.utcnow()
    rows_moved = 0
    already_in_tab = 0
    removed_from_region = 0

    for row in MasterCashbookTransaction.query.filter_by(job_id=job_id).all():
        if not is_partner_transfer_transtype(row.transtype):
            continue

        if row.workbook_partner_transfer:
            already_in_tab += 1
            continue

        if row.workbook_region_id is not None:
            row.workbook_region_id = None
            removed_from_region += 1
        row.workbook_partner_transfer = True
        row.workbook_updated_at = now
        rows_moved += 1

    db.session.commit()

    all_rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    partner_transfer_total = sum(1 for r in all_rows if r.workbook_partner_transfer)
    candidates = sum(
        1 for r in all_rows
        if is_partner_transfer_transtype(r.transtype) and not r.workbook_partner_transfer
    )
    to_be_uploaded = sum(1 for r in all_rows if is_to_be_uploaded_row(r))

    return {
        'rows_moved': rows_moved,
        'already_in_tab': already_in_tab,
        'partner_transfer_total': partner_transfer_total,
        'partner_transfer_candidates_remaining': candidates,
        'removed_from_region_tab': removed_from_region,
        'to_be_uploaded_remaining': to_be_uploaded,
        'master_cashbook_url': f'/prepare/master-cashbook/{job_id}',
    }


def check_sepa_netting(db, job_id, MasterCashbookTransaction):
    """Find paired SEPAREJECT/SEPASUCCESS rows and move both to the SEPA Netting tab."""
    master_count = MasterCashbookTransaction.query.filter_by(job_id=job_id).count()
    if master_count == 0:
        raise ValueError('No master cashbook loaded for this job')

    now = datetime.utcnow()
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    pairs = find_sepa_netting_pairs(rows)
    rows_moved = 0
    pairs_moved = 0
    removed_from_region = 0

    for success, reject in pairs:
        moved_this_pair = False
        for row in (success, reject):
            if row.workbook_sepa_netting:
                continue
            if row.workbook_region_id is not None:
                row.workbook_region_id = None
                removed_from_region += 1
            row.workbook_sepa_netting = True
            row.workbook_updated_at = now
            rows_moved += 1
            moved_this_pair = True
        if moved_this_pair:
            pairs_moved += 1

    db.session.commit()

    all_rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    sepa_netting_total = sum(1 for r in all_rows if r.workbook_sepa_netting)
    sepa_netting_pairs_total = sepa_netting_total // 2
    candidates = count_sepa_netting_candidates(all_rows)
    unmatched_rejects = sum(
        1 for r in all_rows
        if _safe_str(r.transtype).upper() == 'SEPAREJECT' and not r.workbook_sepa_netting
    )
    to_be_uploaded = sum(1 for r in all_rows if is_to_be_uploaded_row(r))

    return {
        'rows_moved': rows_moved,
        'pairs_moved': pairs_moved,
        'sepa_netting_total': sepa_netting_total,
        'sepa_netting_pairs_total': sepa_netting_pairs_total,
        'sepa_netting_candidates_remaining': candidates,
        'unmatched_separeject': unmatched_rejects,
        'removed_from_region_tab': removed_from_region,
        'to_be_uploaded_remaining': to_be_uploaded,
        'master_cashbook_url': f'/prepare/master-cashbook/{job_id}',
    }


def rerun_workbook_region(db, job_id, subsidiary_id, MasterCashbookTransaction, MatchedTransaction, ReconciliationResults):
    """Clear workbook tab and delete all matches for the region."""
    clear_result = clear_workbook_region(db, job_id, subsidiary_id, MasterCashbookTransaction)

    matched_deleted = MatchedTransaction.query.filter_by(
        job_id=job_id, subsidiary_id=subsidiary_id
    ).delete()
    results_deleted = ReconciliationResults.query.filter_by(
        job_id=job_id, subsidiary_id=subsidiary_id
    ).delete()
    db.session.commit()

    return {
        **clear_result,
        'matches_deleted': matched_deleted,
        'reconciliation_results_deleted': results_deleted,
        'redirect_url': f'/reconciliation-process/{job_id}/{subsidiary_id}',
    }


def workbook_summary(job_id, MasterCashbookTransaction):
    """Tab counts and per-region workbook status."""
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).all()
    total = len(rows)
    to_be_uploaded = sum(1 for r in rows if is_to_be_uploaded_row(r))
    uk_bacs_rows = [r for r in rows if r.workbook_uk_bacs]
    uk_bacs_count = len(uk_bacs_rows)
    uk_bacs_candidates = sum(1 for r in rows if is_uk_bacs_transtype(r.transtype) and not r.workbook_uk_bacs)
    partner_transfer_rows = [r for r in rows if r.workbook_partner_transfer]
    partner_transfer_count = len(partner_transfer_rows)
    partner_transfer_candidates = sum(
        1 for r in rows if is_partner_transfer_transtype(r.transtype) and not r.workbook_partner_transfer
    )
    sepa_netting_rows = [r for r in rows if r.workbook_sepa_netting]
    sepa_netting_count = len(sepa_netting_rows)
    sepa_netting_candidates = count_sepa_netting_candidates(rows)
    sepa_netting_pairs_total = sepa_netting_count // 2
    unmatched_separeject = sum(
        1 for r in rows
        if _safe_str(r.transtype).upper() == 'SEPAREJECT' and not r.workbook_sepa_netting
    )
    unusual_matched_total = 0
    by_region = {}
    last_updated = {}
    for sid, label in SUBSIDIARY_LABELS.items():
        region_rows = [r for r in rows if r.workbook_region_id == sid]
        unusual_in_region = sum(
            1 for r in region_rows if is_unusual_transtype_for_stripe_match(r.transtype)[0]
        )
        unusual_matched_total += unusual_in_region
        by_region[label] = {
            'subsidiary_id': sid,
            'tab_key': TAB_KEY_FOR_SUBSIDIARY.get(sid),
            'count': len(region_rows),
            'unusual_matched_count': unusual_in_region,
        }
        updates = [r.workbook_updated_at for r in region_rows if r.workbook_updated_at]
        last_updated[label] = max(updates).isoformat() if updates else None

    tabs = [
        {
            'key': 'master',
            'label': 'Master',
            'count': total,
            'unusual_matched_count': unusual_matched_total,
            'unusual_transtype_count': _count_unusual_transtype(rows),
            'uk_bacs_candidates': uk_bacs_candidates,
            'uk_bacs_total': uk_bacs_count,
            'partner_transfer_candidates': partner_transfer_candidates,
            'partner_transfer_total': partner_transfer_count,
            'sepa_netting_candidates': sepa_netting_candidates,
            'sepa_netting_total': sepa_netting_count,
            'sepa_netting_pairs_total': sepa_netting_pairs_total,
            'unmatched_separeject': unmatched_separeject,
        },
        {
            'key': 'to_be_uploaded',
            'label': 'To be Uploaded',
            'count': to_be_uploaded,
            'unusual_matched_count': 0,
            'unusual_transtype_count': _count_unusual_transtype(
                [r for r in rows if is_to_be_uploaded_row(r)]
            ),
        },
    ]
    if uk_bacs_count > 0:
        tabs.append({
            'key': 'uk_bacs',
            'label': 'UK Bacs',
            'count': uk_bacs_count,
            'unusual_matched_count': 0,
            'unusual_transtype_count': 0,
        })
    if partner_transfer_count > 0:
        tabs.append({
            'key': 'partner_transfer',
            'label': 'Partner Transfer',
            'count': partner_transfer_count,
            'unusual_matched_count': 0,
            'unusual_transtype_count': 0,
        })
    if sepa_netting_count > 0:
        tabs.append({
            'key': 'sepa_netting',
            'label': 'SEPA Netting',
            'count': sepa_netting_count,
            'unusual_matched_count': 0,
            'unusual_transtype_count': 0,
        })
    for sid, label in SUBSIDIARY_LABELS.items():
        count = by_region[label]['count']
        if count > 0:
            region_rows = [r for r in rows if r.workbook_region_id == sid]
            tabs.append({
                'key': TAB_KEY_FOR_SUBSIDIARY[sid],
                'label': f'{label} Stripe Import',
                'count': count,
                'subsidiary_id': sid,
                'unusual_matched_count': by_region[label]['unusual_matched_count'],
                'unusual_transtype_count': _count_unusual_transtype(region_rows),
            })

    return {
        'total_rows': total,
        'to_be_uploaded': to_be_uploaded,
        'uk_bacs_total': uk_bacs_count,
        'uk_bacs_candidates': uk_bacs_candidates,
        'partner_transfer_total': partner_transfer_count,
        'partner_transfer_candidates': partner_transfer_candidates,
        'sepa_netting_total': sepa_netting_count,
        'sepa_netting_pairs_total': sepa_netting_pairs_total,
        'sepa_netting_candidates': sepa_netting_candidates,
        'unmatched_separeject': unmatched_separeject,
        'unusual_matched_total': unusual_matched_total,
        'by_region': by_region,
        'last_updated_by_region': last_updated,
        'tabs': tabs,
    }


def build_workbook_partitions(job_id, MasterCashbookTransaction):
    """Return row lists keyed by workbook tab."""
    rows = MasterCashbookTransaction.query.filter_by(job_id=job_id).order_by(MasterCashbookTransaction.id).all()
    partitions = {
        'master': rows,
        'to_be_uploaded': [r for r in rows if is_to_be_uploaded_row(r)],
        'uk_bacs': [r for r in rows if r.workbook_uk_bacs],
        'partner_transfer': [r for r in rows if r.workbook_partner_transfer],
        'sepa_netting': [r for r in rows if r.workbook_sepa_netting],
        'regions': {sid: [r for r in rows if r.workbook_region_id == sid] for sid in SUBSIDIARY_LABELS},
    }
    return partitions


_SORTABLE_COLS = {
    'payment_date', 'client_id', 'invoice_number', 'billing_entity',
    'currency', 'amount', 'account', 'location', 'transtype',
    'payment_hash', 'reasoncode',
}


def _workbook_order_expr(MasterCashbookTransaction, order_col, sort_dir='asc'):
    """Return SQLAlchemy order expression; payment_date uses chronological DD/MM/YYYY parse."""
    from sqlalchemy import func

    attr = getattr(MasterCashbookTransaction, order_col, MasterCashbookTransaction.id)
    if order_col == 'payment_date':
        expr = func.to_date(MasterCashbookTransaction.payment_date, 'DD/MM/YYYY')
    else:
        expr = attr
    return expr.desc() if sort_dir == 'desc' else expr.asc()


def _workbook_tab_base_query(MasterCashbookTransaction, job_id, tab_key, unusual_only=False):
    """Base query scoped to a workbook tab."""
    tab_key = tab_key.lower()
    q = MasterCashbookTransaction.query.filter_by(job_id=job_id)

    if tab_key == 'master':
        pass
    elif tab_key == 'to_be_uploaded':
        q = _workbook_to_be_uploaded_query(q, MasterCashbookTransaction)
    elif tab_key == 'uk_bacs':
        q = q.filter(MasterCashbookTransaction.workbook_uk_bacs.is_(True))
    elif tab_key == 'partner_transfer':
        q = q.filter(MasterCashbookTransaction.workbook_partner_transfer.is_(True))
    elif tab_key == 'sepa_netting':
        q = q.filter(MasterCashbookTransaction.workbook_sepa_netting.is_(True))
    elif tab_key in WORKBOOK_TAB_KEYS and isinstance(WORKBOOK_TAB_KEYS[tab_key], int):
        q = q.filter_by(workbook_region_id=WORKBOOK_TAB_KEYS[tab_key])
    else:
        raise ValueError(f'Unknown tab key: {tab_key}')

    if unusual_only:
        q = _apply_unusual_transtype_query(q, MasterCashbookTransaction)
    return q


def _apply_workbook_table_filters(
    q, MasterCashbookTransaction, col_filters=None, global_search=None, skip_col=None,
):
    """Apply column + global search filters; skip_col excludes one column (for distinct lookups)."""
    from sqlalchemy import or_, cast, String, func

    if col_filters:
        for col, val in col_filters.items():
            if col == skip_col or col not in _SORTABLE_COLS:
                continue
            if isinstance(val, (list, tuple)):
                values = [str(v).strip() for v in val if v is not None and str(v).strip()]
            else:
                values = [str(val).strip()] if val is not None and str(val).strip() else []
            if not values:
                continue
            attr = getattr(MasterCashbookTransaction, col, None)
            if attr is None:
                continue
            lowered = [v.lower() for v in values]
            if len(lowered) == 1:
                q = q.filter(func.lower(cast(attr, String)) == lowered[0])
            else:
                q = q.filter(func.lower(cast(attr, String)).in_(lowered))

    if global_search and global_search.strip():
        term = global_search.strip().lower()
        search_cols = ['billing_entity', 'transtype', 'location', 'account',
                       'currency', 'payment_hash', 'invoice_number']
        q = q.filter(or_(
            *[func.lower(cast(getattr(MasterCashbookTransaction, c), String)).contains(term)
              for c in search_cols]
        ))
    return q


def workbook_column_distinct_values(
    MasterCashbookTransaction, job_id, tab_key, column,
    unusual_only=False, col_filters=None, global_search=None, limit=300,
):
    """Distinct values for a column within a tab (respecting other active filters)."""
    from sqlalchemy import cast, String, func

    column = column.lower()
    if column not in _SORTABLE_COLS:
        raise ValueError(f'Unknown column: {column}')

    q = _workbook_tab_base_query(MasterCashbookTransaction, job_id, tab_key, unusual_only)
    q = _apply_workbook_table_filters(
        q, MasterCashbookTransaction,
        col_filters=col_filters, global_search=global_search, skip_col=column,
    )
    attr = getattr(MasterCashbookTransaction, column)
    rows = (
        q.with_entities(cast(attr, String).label('v'))
        .filter(attr.isnot(None))
        .distinct()
        .order_by(cast(attr, String).asc())
        .limit(limit)
        .all()
    )
    values = []
    for (raw,) in rows:
        s = str(raw).strip() if raw is not None else ''
        if s and s.lower() != 'nan':
            values.append(s)
    return values


def master_rows_for_workbook_tab(
    MasterCashbookTransaction, job_id, tab_key,
    page=1, per_page=100, unusual_only=False,
    sort_by=None, sort_dir='asc',
    col_filters=None, global_search=None,
):
    """Paginated rows for a workbook tab with server-side sort / filter / search."""
    tab_key = tab_key.lower()
    q = _workbook_tab_base_query(MasterCashbookTransaction, job_id, tab_key, unusual_only)
    q = _apply_workbook_table_filters(q, MasterCashbookTransaction, col_filters, global_search)

    total = q.count()

    # ordering
    order_col = sort_by if sort_by in _SORTABLE_COLS else 'id'
    if order_col == 'id':
        q = q.order_by(MasterCashbookTransaction.id.asc() if sort_dir != 'desc' else MasterCashbookTransaction.id.desc())
    else:
        order_expr = _workbook_order_expr(MasterCashbookTransaction, order_col, sort_dir)
        q = q.order_by(order_expr, MasterCashbookTransaction.id.asc())

    rows = q.offset((page - 1) * per_page).limit(per_page).all()
    return rows, total


def workbook_tab_stats(rows):
    """Compute dashboard stats for a list of MasterCashbookTransaction rows."""
    from collections import Counter
    total = len(rows)
    total_amount = sum(float(r.amount or 0) for r in rows)
    by_transtype = Counter()
    by_currency = Counter()
    for r in rows:
        tt = _safe_str(r.transtype).upper() or 'UNKNOWN'
        by_transtype[tt] += 1
        cu = _safe_str(r.currency).upper() or 'UNKNOWN'
        by_currency[cu] += 1

    # top-5 transtypes by count, rest collapsed into 'Other'
    top_tt = by_transtype.most_common(8)
    by_currency_list = [{'currency': k, 'count': v} for k, v in by_currency.most_common()]

    return {
        'total_rows': total,
        'total_amount': round(total_amount, 2),
        'by_transtype': [{'transtype': k, 'count': v} for k, v in top_tt],
        'by_currency': by_currency_list,
    }


def _apply_unusual_match_highlight(writer, sheet_name, rows):
    """Orange highlight for matched rows with non-standard transtype (e.g. MANUAL)."""
    from openpyxl.styles import Font, PatternFill

    ws = writer.sheets.get(sheet_name)
    if not ws:
        return
    fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    font = Font(bold=True, color='1A1A1A')
    for idx, row in enumerate(rows, start=2):
        if not row.workbook_region_id:
            continue
        unusual, _ = is_unusual_transtype_for_stripe_match(row.transtype)
        if not unusual:
            continue
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=idx, column=col)
            cell.fill = fill
            cell.font = font


def export_workbook_excel(job_id, MasterCashbookTransaction):
    """Multi-sheet workbook Excel export."""
    from io import BytesIO

    partitions = build_workbook_partitions(job_id, MasterCashbookTransaction)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        master_df = master_to_export_dataframe(partitions['master'])
        master_df.to_excel(writer, sheet_name='Master', index=False)

        tb_df = master_to_export_dataframe(partitions['to_be_uploaded'])
        tb_df.to_excel(writer, sheet_name='To be Uploaded', index=False)

        uk_bacs_rows = partitions.get('uk_bacs') or []
        if uk_bacs_rows:
            uk_df = master_to_export_dataframe(uk_bacs_rows)
            uk_df.to_excel(writer, sheet_name='UK Bacs', index=False)

        partner_transfer_rows = partitions.get('partner_transfer') or []
        if partner_transfer_rows:
            pt_df = master_to_export_dataframe(partner_transfer_rows)
            pt_df.to_excel(writer, sheet_name='Partner Transfer', index=False)

        sepa_netting_rows = partitions.get('sepa_netting') or []
        if sepa_netting_rows:
            sn_df = master_to_export_dataframe(sepa_netting_rows)
            sn_df.to_excel(writer, sheet_name='SEPA Netting', index=False)

        for sid, label in SUBSIDIARY_LABELS.items():
            region_rows = partitions['regions'].get(sid) or []
            if not region_rows:
                continue
            sheet = f'{label} Stripe Import'
            df = master_to_export_dataframe(region_rows)
            df.to_excel(writer, sheet_name=sheet, index=False)
            _apply_unusual_match_highlight(writer, sheet, region_rows)

    output.seek(0)
    return output

